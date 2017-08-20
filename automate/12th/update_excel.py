#! python3
# update the price of Celery Garlic Lemon

import openpyxl
import os

wb = openpyxl.load_workbook('res/produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# 数据字典
price_updates = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# 修改价格
for row_index in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row_index, column=1).value
    if produce_name in price_updates:
        sheet.cell(row=row_index, column=2).value = price_updates[produce_name]

# 保存更新后的表格
wb.save('res/update_produceSales.xlsx')
# 删除原来的表格
os.remove('res/produceSales.xlsx')
