#! python
# insert blank rows
import openpyxl
import os


def insert_rows(name, index=3, rows=2):
    # 拼接res+filename
    filename = os.path.join('res', name)
    # 判断文件是否存在
    if not os.path.exists(filename):
        return
    # 打开指定文件
    wb = openpyxl.load_workbook(filename)
    # 获取使用的表格
    active_sheet = wb.get_active_sheet()
    cells_data = {}
    # 读取从index开始到最后的所有行的数据
    for row_index in range(index, active_sheet.max_row + 1):
        for column_index in range(1, active_sheet.max_column + 1):
            # if str(row_index) not in cells_data.keys():
            cells_data.setdefault(str(row_index), [])
            cells_data[str(row_index)].append(active_sheet.cell(row=row_index, column=column_index).value)
    # 清空index后的所有数据
    for row_index in range(index, active_sheet.max_row + 1):
        for column_index in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=row_index, column=column_index).value = ''
    # 将数据重新写入到excel中 从 index + rows 的位置开始
    for row_index in range(index + rows, active_sheet.max_row + rows + 1):
        lst = cells_data[str(row_index - rows)]
        for column_index in range(1, len(lst) + 1):
            active_sheet.cell(row=row_index, column=column_index).value = lst[column_index - 1]
    # 保存修改后的excle文档
    wb.save(filename)

insert_rows('example.xlsx', index=3, rows=2)
