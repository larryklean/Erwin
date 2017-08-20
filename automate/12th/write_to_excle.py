import openpyxl

# create new file
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
print(wb.get_active_sheet().title)

# create new table
wb.create_sheet()
print(wb.get_sheet_names())

# cretae new table with index and title
wb.create_sheet(index=0, title='First Sheet')
print(wb.get_sheet_names())

wb.create_sheet(index=0, title='Second Sheet')
print(wb.get_sheet_names())

# delete table
wb.remove(wb.get_sheet_by_name('Second Sheet'))
print(wb.get_sheet_names())
