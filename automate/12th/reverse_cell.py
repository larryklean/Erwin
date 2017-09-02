#! python3
# reverse data in cells

import openpyxl
import os


def reverse_data(filename):
    # 判断文件是否存在
    if not os.path.exists(os.path.abspath(os.path.join('res', filename))):
        print('File not exists!')
        return
    # 获取excle文件对象
    wb = openpyxl.load_workbook(os.path.join('res', filename))
    # 获取其中的表格
    active_sheet = wb.get_active_sheet()
    # 翻转表格内容
    sheet_data = []
    for row_index in range(1, active_sheet.max_row + 1):
        cell_lst = []
        for column_index in range(1, active_sheet.max_column + 1):
            cell_value = active_sheet.cell(row=row_index, column=column_index).value
            cell_lst.append(cell_value)
        sheet_data.append(cell_lst)
    # 清空excel表格中是数据
    for row_index in range(1, active_sheet.max_row + 1):
        for column_index in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=row_index, column=column_index).value = ''
    # 反转数据
    for i in range(0, len(sheet_data)):
        for j in range(0, len(sheet_data[i])):
            active_sheet.cell(row=j + 1, column=i + 1).value = sheet_data[i][j]
    # 保存修改后的excel文件
    wb.save(os.path.join('res', filename))


reverse_data('example.xlsx')
