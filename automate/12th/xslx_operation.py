import openpyxl
import os

file_path = os.path.join('res', 'example.xlsx')
# 加载
wb = openpyxl.load_workbook(file_path)
print(type(wb))

# 获取所有的表格
print(wb.get_sheet_names())

# 获取某个表格对象
sheet = wb.get_sheet_by_name('Sheet1')
print(type(sheet))

# 获取表格的标题
print(sheet.title)

# 获取活动表 即打开文档时显示的表
print(wb.get_active_sheet())

# 获取单元格对象
cell = sheet['A1']
print(cell.value)

cell = sheet['B1']
print(cell.value)

# 行、列、值
print('Row ' + str(cell.row) + ' Column ' + str(cell.column) + ' Value is ' + cell.value)

# cell()方法
cell = sheet.cell(row=1, column=2)
print(cell)
print(cell.value)

for index in range(1, 8):
    print(index, sheet.cell(row=index, column=2).value)

# 列字母和数字之间的转换
# letter = get_column_letter(1)
# print(letter)

# 获取行和列
wb2 = openpyxl.load_workbook('res/example.xlsx')
sheet2 = wb2.get_sheet_by_name('Sheet1')
# print(tuple(sheet2['A1':'C3']))
for cell_objs in sheet2['A1':'C3']:
    for cell_obj in cell_objs:
        print(cell_obj.value)
    print('---End Row---')
