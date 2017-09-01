#! python3
# make a n x n table of number
import openpyxl


def make_excel(n):
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = "nn table"
    for i in range(1, n + 1):
        sheet.cell(row=1, column=i + 1).value = i
        sheet.cell(row=i + 1, column=1).value = i
    for row_index in range(2, sheet.max_row + 1):
        for column_index in range(2, sheet.max_column + 1):
            row_value = sheet.cell(row=row_index, column=1).value
            column_value = sheet.cell(row=1, column=column_index).value
            sheet.cell(row=row_index, column=column_index).value = row_value * column_value
    wb.save('res/nn_table.xlsx')


make_excel(5)
